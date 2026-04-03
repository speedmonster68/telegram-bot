import os
import tempfile
from pathlib import Path
from typing import Dict

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
from telegram import Update
from telegram.ext import Application, CommandHandler, ContextTypes, MessageHandler, filters

BOT_TOKEN = "8294093241:AAHKUedg20GgUOpMxAhKFUP3IocVhCKer0k"

TEMPLATE_PATH = Path("template.png")
FONT_PATH = Path("SBSansDisplay-Regular.ttf")

COLOR_DARK = (58, 68, 79, 255)
COLOR_GREEN = (141, 194, 44, 255)
COLOR_BLUE = (79, 192, 242, 255)

FONT_DATE_SIZE = 29
FONT_VALUE_SIZE = 33
FONT_CHANGE_SIZE = 33

COORDS = {
    "date": (1106, 49),

    "imoex_value": (341, 225),
    "imoex_change": (332, 320),

    "rts_value": (560, 225),
    "rts_change": (557, 320),

    "brent_value": (772, 225),
    "brent_change": (746, 320),

    "gold_value": (994, 225),
    "gold_change": (978, 320),

    "cny_value": (329, 585),
    "cny_change": (334, 680),

    "usd_value": (558, 585),
    "usd_change": (555, 680),

    "eur_value": (767, 585),
    "eur_change": (766, 680),

    "aed_value": (986, 585),
    "aed_change": (987, 680),

    "bond2_value": (332, 945),
    "bond2_change": (332, 1041),

    "bond5_value": (556, 945),
    "bond5_change": (577, 1041),

    "bond10_value": (768, 945),
    "bond10_change": (789, 1041),

    "bond20_value": (988, 945),
    "bond20_change": (1014, 1041),
}

CHANGE_COLORS = {
    "imoex_change": COLOR_BLUE,
    "rts_change": COLOR_GREEN,
    "brent_change": COLOR_GREEN,
    "gold_change": COLOR_BLUE,

    "cny_change": COLOR_BLUE,
    "usd_change": COLOR_BLUE,
    "eur_change": COLOR_BLUE,
    "aed_change": COLOR_BLUE,

    "bond2_change": COLOR_DARK,
    "bond5_change": COLOR_DARK,
    "bond10_change": COLOR_DARK,
    "bond20_change": COLOR_DARK,
}


def read_excel(xlsx_path: str) -> Dict[str, str]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    data: Dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        key = row[0]
        value = row[1] if len(row) > 1 else ""
        if key:
            data[str(key).strip()] = "" if value is None else str(value).strip()
    return data


def draw_center(draw: ImageDraw.ImageDraw, xy, text: str, font, fill) -> None:
    bbox = draw.textbbox((0, 0), text, font=font)
    w = bbox[2] - bbox[0]
    x = int(xy[0] - w / 2)
    y = int(xy[1])
    draw.text((x, y), text, font=font, fill=fill)


def render(template_path: Path, output_path: str, data: Dict[str, str]) -> None:
    image = Image.open(template_path).convert("RGBA")
    draw = ImageDraw.Draw(image)

    font_date = ImageFont.truetype(str(FONT_PATH), FONT_DATE_SIZE)
    font_value = ImageFont.truetype(str(FONT_PATH), FONT_VALUE_SIZE)
    font_change = ImageFont.truetype(str(FONT_PATH), FONT_CHANGE_SIZE)

    draw_center(draw, COORDS["date"], data.get("date", ""), font_date, COLOR_DARK)

    value_keys = [
        "imoex_value", "rts_value", "brent_value", "gold_value",
        "cny_value", "usd_value", "eur_value", "aed_value",
        "bond2_value", "bond5_value", "bond10_value", "bond20_value",
    ]
    for key in value_keys:
        draw_center(draw, COORDS[key], data.get(key, ""), font_value, COLOR_DARK)

    change_keys = [
        "imoex_change", "rts_change", "brent_change", "gold_change",
        "cny_change", "usd_change", "eur_change", "aed_change",
        "bond2_change", "bond5_change", "bond10_change", "bond20_change",
    ]
    for key in change_keys:
        draw_center(draw, COORDS[key], data.get(key, ""), font_change, CHANGE_COLORS[key])

    image.save(output_path)


async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message:
        await update.message.reply_text("Отправь Excel файл .xlsx")


async def handle(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.message or not update.message.document:
        return

    doc = update.message.document

    if not doc.file_name or not doc.file_name.lower().endswith(".xlsx"):
        await update.message.reply_text("Нужен .xlsx файл")
        return

    with tempfile.TemporaryDirectory() as tmp:
        file_path = os.path.join(tmp, doc.file_name)
        result_path = os.path.join(tmp, "result.png")

        tg_file = await context.bot.get_file(doc.file_id)
        await tg_file.download_to_drive(file_path)

        data = read_excel(file_path)
        render(TEMPLATE_PATH, result_path, data)

        with open(result_path, "rb") as img:
            await update.message.reply_photo(img)


def main():
    if not TEMPLATE_PATH.exists():
        raise Exception("Нет template.png")

    if not FONT_PATH.exists():
        raise Exception(f"Нет шрифта {FONT_PATH}")

    app = (
        Application.builder()
        .token(BOT_TOKEN)
        .connect_timeout(30)
        .read_timeout(30)
        .write_timeout(30)
        .pool_timeout(30)
        .build()
    )

    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.Document.ALL, handle))

    print("Бот запущен 🚀")
    app.run_polling()


if __name__ == "__main__":
    main()
